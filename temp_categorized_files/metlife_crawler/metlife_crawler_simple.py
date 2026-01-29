"""
메트라이프 크롤러 - 완전 자동화 버전
메트라이프 변액보험 기준가 데이터를 자동으로 수집하여 엑셀로 저장합니다.
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import time
from datetime import datetime
import sys
import os
import re

def setup_driver():
    """Chrome 드라이버 설정"""
    options = webdriver.ChromeOptions()
    options.add_argument('--lang=ko-KR')
    options.add_argument('--disable-blink-features=AutomationControlled')
    # 페이지 로딩 전략 최적화
    options.page_load_strategy = 'normal'
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.maximize_window()
    return driver

def main():
    print("=" * 80)
    print("메트라이프 변액보험 기준가 크롤러 (완전 자동화)")
    print("=" * 80)
    print("\n📸 디버깅 모드: 각 체크박스 시도마다 스크린샷을 저장합니다.")
    print("   → 파일명: checkbox_1_result.png, checkbox_2_result.png, ...")
    print("   → 위치: 현재 폴더")
    print("\n" + "=" * 80)
    
    driver = setup_driver()
    wait = WebDriverWait(driver, 15)
    
    try:
        url = "https://brand.metlife.co.kr/pn/paReal/retrieveVrinsPaBprcPcndList.do"
        
        print(f"\n[1단계] 페이지 접속: {url}")
        driver.get(url)
        time.sleep(3)
        
        print("\n[2단계] 판매중지상품 탭 클릭 중...")
        try:
            discontinued_tab = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '판매중지상품')]")))
            driver.execute_script("arguments[0].click();", discontinued_tab)
            time.sleep(2)
        except:
            print("   ⚠ 탭 클릭 실패 (이미 열려있을 수 있습니다.)")
        
        print("\n[3단계] 상품 선택 시도...")
        try:
            select_box = wait.until(EC.presence_of_element_located((By.XPATH, "//select")))
            options = select_box.find_elements(By.TAG_NAME, "option")
            
            product_found = False
            for option in options:
                full_text = option.text.strip()
                compare_text = full_text.replace(" ", "").lower()
                if "myfund" in compare_text and "유니버셜" in compare_text:
                    print(f"   ✅ 대상 상품 발견 및 선택: {full_text}")
                    option.click()
                    product_found = True
                    break
            
            if not product_found:
                print("   ❌ 대상 상품('My Fund' & '유니버셜')을 목록에서 찾을 수 없습니다.")
        except Exception as e:
            print(f"   ⚠ 상품 선택 중 오류: {e}")

        print("\n[4단계] 검색 버튼 클릭 시도...")
        try:
            search_btn_xpath = (
                "//button[contains(., '검색')] | "
                "//a[contains(., '검색')] | "
                "//input[@value='검색'] | "
                "//span[contains(text(), '검색')]/parent::*"
            )
            search_btns = driver.find_elements(By.XPATH, search_btn_xpath)
            if search_btns:
                driver.execute_script("arguments[0].click();", search_btns[0])
                print("   ✓ 검색 클릭 시도 완료.")
                time.sleep(2)
            else:
                print("   ⚠ 검색 버튼을 찾지 못했습니다.")
        except:
            pass

        print("\n[5단계] [보기] 버튼 자동 클릭 시도...")
        try:
            # 방법 1: 테이블에서 My Fund & 유니버셜 행의 보기 링크 찾기
            view_link_found = False
            
            # 모든 테이블 행 검색
            table_rows = driver.find_elements(By.XPATH, "//table//tr")
            for row in table_rows:
                row_text = row.text.replace(" ", "").lower()
                if "myfund" in row_text and "유니버셜" in row_text:
                    # 해당 행에서 "보기" 링크 찾기
                    view_links = row.find_elements(By.XPATH, ".//a[contains(text(), '보기') or contains(@onclick, '보기')]")
                    if view_links:
                        print(f"   ✅ [보기] 링크 발견: {row.text[:50]}...")
                        driver.execute_script("arguments[0].scrollIntoView(true);", view_links[0])
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", view_links[0])
                        print("   ✓ [보기] 버튼 클릭 완료")
                        view_link_found = True
                        break
            
            if not view_link_found:
                # 방법 2: 모든 "보기" 링크 중 첫 번째 시도
                all_view_links = driver.find_elements(By.XPATH, "//a[contains(text(), '보기')]")
                if all_view_links:
                    print(f"   ⚠ 특정 행을 찾지 못해 첫 번째 [보기] 클릭 시도 (총 {len(all_view_links)}개)")
                    driver.execute_script("arguments[0].click();", all_view_links[0])
                    print("   ✓ [보기] 버튼 클릭 완료")
                    view_link_found = True
            
            if not view_link_found:
                raise Exception("[보기] 버튼을 찾을 수 없습니다.")
            
            time.sleep(3)  # 페이지 전환 대기
            
        except Exception as e:
            print(f"   ❌ [보기] 버튼 자동 클릭 실패: {e}")
            driver.save_screenshot("error_view_button.png")
            print("   스크린샷 저장: error_view_button.png")
            print("\n   ⚠ 수동 작업이 필요합니다:")
            print("   1. 브라우저에서 '무배당 My Fund 변액유니버셜보험' 줄의 [보기]를 클릭하세요")
            print("   2. 페이지가 바뀌면 이 창으로 돌아와서 [Enter]를 눌러주세요")
            input("\n   >>> [보기] 클릭 완료? [Enter] 입력: ")
            time.sleep(2)

        print("\n[6단계] 기준가 현황 페이지 로딩 완료")
        time.sleep(1)
        
        print("\n[7단계] 혼합성장형 펀드 찾기 (스마트 검색)")
        print("   → 먼저 모든 체크박스 라벨을 읽고, 혼합성장형부터 시도합니다...")
        
        mixed_found = False
        found_fund_name = ""  # 찾은 펀드명 저장용
        time.sleep(2)  # 페이지 안정화
        
        # 1단계: 모든 체크박스와 라벨 미리 읽기
        all_checkboxes = driver.find_elements(By.XPATH, "//input[@type='checkbox']")
        print(f"   → 총 {len(all_checkboxes)}개의 체크박스 발견")
        
        # 체크박스 정보 저장 (인덱스, 체크박스 요소, 라벨)
        checkbox_info = []
        
        print("\n   [체크박스 라벨 스캔 중...]")
        for idx, checkbox in enumerate(all_checkboxes, 1):
            checkbox_label = f"체크박스{idx}"
            try:
                # 방법 1: label 태그 찾기
                checkbox_id = checkbox.get_attribute('id')
                if checkbox_id:
                    try:
                        label = driver.find_element(By.XPATH, f"//label[@for='{checkbox_id}']")
                        if label.text.strip():
                            checkbox_label = label.text.strip()
                    except:
                        pass
                
                # 방법 2: 부모의 텍스트에서 추출
                if checkbox_label == f"체크박스{idx}":
                    try:
                        parent = checkbox.find_element(By.XPATH, "..")
                        parent_text = parent.text.strip()
                        lines = [line.strip() for line in parent_text.split('\n') if line.strip()]
                        # 체크박스 옆의 짧은 텍스트만 (다른 체크박스 라벨 제외)
                        for line in lines:
                            if len(line) < 30 and '형' in line:
                                checkbox_label = line
                                break
                    except:
                        pass
                
                # 방법 3: JavaScript로 라벨 찾기
                if checkbox_label == f"체크박스{idx}":
                    try:
                        label_text = driver.execute_script("""
                            var cb = arguments[0];
                            var label = cb.nextElementSibling;
                            if (label && label.textContent) {
                                return label.textContent.trim();
                            }
                            return '';
                        """, checkbox)
                        if label_text and len(label_text) < 30:
                            checkbox_label = label_text
                    except:
                        pass
                        
            except Exception as e:
                pass
            
            checkbox_info.append({
                'index': idx,
                'element': checkbox,
                'label': checkbox_label
            })
            print(f"      {idx}. {checkbox_label}")
        
        # 2단계: "혼합성장형"이 있으면 맨 앞으로 이동
        mixed_index = -1
        for i, info in enumerate(checkbox_info):
            if "혼합성장형" in info['label']:
                mixed_index = i
                break
        
        if mixed_index >= 0:
            print(f"\n   ✨ '혼합성장형' 발견! (위치: {mixed_index + 1}) → 먼저 시도합니다.")
            # 혼합성장형을 맨 앞으로 이동
            mixed_info = checkbox_info.pop(mixed_index)
            checkbox_info.insert(0, mixed_info)
        else:
            print(f"\n   ⚠ 라벨에서 '혼합성장형'을 못 찾았습니다. 순서대로 시도합니다.")
        
        # 3단계: 재정렬된 순서로 체크박스 시도
        print(f"\n   [시도 순서]")
        for i, info in enumerate(checkbox_info, 1):
            print(f"      {i}. {info['label']} (원래 {info['index']}번)")
        
        print("\n" + "="*80)
        
        for attempt_num, info in enumerate(checkbox_info, 1):
            idx = info['index']
            checkbox = info['element']
            checkbox_label = info['label']
            
            # 시도 시작 헤더 출력
            print(f"\n   {'='*70}")
            print(f"   [시도 {attempt_num}/{len(checkbox_info)}] 체크박스 {idx}: {checkbox_label}")
            print(f"   {'='*70}")
            
            try:
                # 1. 체크박스 재탐색 (stale element 방지)
                try:
                    # 페이지가 변경되었을 수 있으므로 체크박스를 다시 찾기
                    fresh_checkboxes = driver.find_elements(By.XPATH, "//input[@type='checkbox']")
                    if idx <= len(fresh_checkboxes):
                        checkbox = fresh_checkboxes[idx - 1]
                except:
                    pass
                
                # 2. 모든 체크박스 해제
                all_cbs = driver.find_elements(By.XPATH, "//input[@type='checkbox']")
                for cb in all_cbs:
                    try:
                        if cb.is_selected():
                            driver.execute_script("arguments[0].click();", cb)
                    except:
                        pass
                time.sleep(0.5)
                
                # 3. 현재 체크박스 선택
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", checkbox)
                time.sleep(0.5)
                print(f"      ✓ 체크박스 선택 완료 ({checkbox_label})")
                
                # 4. 3개월 선택 (필수!)
                period_set = False
                try:
                    # 방법 1: 3개월 버튼 클릭 시도
                    three_m_xpath = (
                        "//a[text()='3개월'] | "
                        "//a[contains(text(), '3개월')] | "
                        "//span[text()='3개월'] | "
                        "//button[contains(text(), '3개월')] | "
                        "//*[contains(@class, 'month') and contains(text(), '3')]"
                    )
                    three_m_elements = driver.find_elements(By.XPATH, three_m_xpath)
                    print(f"      → {len(three_m_elements)}개의 '3개월' 관련 요소 발견")
                    
                    for elem in three_m_elements:
                        try:
                            if elem.is_displayed():
                                elem_text = elem.text.strip()
                                print(f"      → 시도: '{elem_text}'")
                                if "3개월" in elem_text or elem_text == "3개월":
                                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elem)
                                    time.sleep(0.3)
                                    driver.execute_script("arguments[0].click();", elem)
                                    time.sleep(0.5)
                                    print(f"      ✓ 3개월 버튼 클릭 완료!")
                                    period_set = True
                                    break
                        except:
                            continue
                    
                    # 방법 2: 3개월 버튼 못 찾으면 날짜 직접 입력
                    if not period_set:
                        print(f"      ⚠ 3개월 버튼을 못 찾았습니다. 날짜를 직접 입력합니다...")
                        
                        today = datetime.now().strftime('%Y%m%d')
                        start_date = '20250101'
                        
                        print(f"      → 기간: {start_date} ~ {today}")
                        
                        # 시작일/종료일 입력 필드 찾기 (다양한 방법 시도)
                        date_inputs = []
                        
                        # 시도 1: name이나 id에 date 포함
                        date_inputs = driver.find_elements(By.XPATH, "//input[@type='text' and (contains(@name, 'date') or contains(@id, 'date') or contains(@class, 'date'))]")
                        
                        # 시도 2: maxlength="8" (YYYYMMDD 형식)
                        if len(date_inputs) < 2:
                            date_inputs = driver.find_elements(By.XPATH, "//input[@type='text' and @maxlength='8']")
                        
                        # 시도 3: 검색기간 근처의 input
                        if len(date_inputs) < 2:
                            date_inputs = driver.find_elements(By.XPATH, "//*[contains(text(), '검색기간')]//following::input[@type='text']")
                        
                        # 시도 4: 모든 text input 중 보이는 것만
                        if len(date_inputs) < 2:
                            all_inputs = driver.find_elements(By.XPATH, "//input[@type='text']")
                            date_inputs = [inp for inp in all_inputs if inp.is_displayed()]
                        
                        print(f"      → {len(date_inputs)}개의 날짜 입력 필드 발견")
                        
                        if len(date_inputs) >= 2:
                            try:
                                # 첫 번째 날짜 필드: 시작일
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", date_inputs[0])
                                time.sleep(0.3)
                                driver.execute_script("arguments[0].value = '';", date_inputs[0])  # JS로 clear
                                date_inputs[0].click()
                                time.sleep(0.2)
                                date_inputs[0].send_keys(start_date)
                                print(f"      ✓ 시작일 입력: {start_date}")
                                
                                # 두 번째 날짜 필드: 종료일
                                time.sleep(0.3)
                                driver.execute_script("arguments[0].value = '';", date_inputs[1])  # JS로 clear
                                date_inputs[1].click()
                                time.sleep(0.2)
                                date_inputs[1].send_keys(today)
                                print(f"      ✓ 종료일 입력: {today}")
                                
                                # 입력 완료 후 포커스 이동 (입력 확정)
                                time.sleep(0.3)
                                driver.execute_script("arguments[0].blur();", date_inputs[1])
                                
                                period_set = True
                            except Exception as date_err:
                                print(f"      ⚠ 날짜 입력 중 오류: {date_err}")
                        else:
                            print(f"      ⚠ 날짜 입력 필드를 찾지 못했습니다. (발견: {len(date_inputs)}개)")
                    
                    if not period_set:
                        print(f"      ⚠ 기간 설정 실패. 이 체크박스는 건너뜁니다.")
                        continue  # 다음 체크박스로
                        
                except Exception as e:
                    print(f"      ⚠ 기간 설정 실패: {e}")
                    continue  # 다음 체크박스로
                
                # 5. 검색 버튼 클릭
                try:
                    search_xpath = "//button[contains(., '검색')] | //a[contains(., '검색')] | //input[@value='검색']"
                    search_btns = driver.find_elements(By.XPATH, search_xpath)
                    for btn in search_btns:
                        if btn.is_displayed():
                            driver.execute_script("arguments[0].click();", btn)
                            print(f"      ✓ 검색 버튼 클릭 완료")
                            time.sleep(4)  # 결과 로딩 대기 (3초 → 4초)
                            break
                except Exception as e:
                    print(f"      ⚠ 검색 버튼 클릭 실패: {e}")
                    continue
                
                # 6. 테이블 상단의 실제 펀드명 읽기 (테이블 주변에서만!)
                print(f"\n      [결과 분석]")
                actual_fund_name = "찾을 수 없음"
                try:
                    # 먼저 테이블 찾기
                    tables = driver.find_elements(By.TAG_NAME, "table")
                    
                    if tables:
                        main_table = tables[0]  # 첫 번째 테이블
                        
                        # 방법 1: 테이블 바로 위의 요소들에서 펀드명 찾기
                        try:
                            # 테이블의 이전 형제 요소들 확인
                            preceding_elements = main_table.find_elements(By.XPATH, "./preceding-sibling::*[position()<=3]")
                            for elem in reversed(preceding_elements):  # 가까운 것부터
                                elem_text = elem.text.strip()
                                if "형" in elem_text and len(elem_text) < 30:
                                    # 펀드 유형 키워드 확인
                                    fund_keywords = ["혼합성장형", "글로벌주식형", "주식형", "채권형", "안정형", "MMF형", "성장형", "혼합안정형"]
                                    for keyword in fund_keywords:
                                        if keyword in elem_text:
                                            actual_fund_name = keyword
                                            print(f"      → 테이블 위 요소에서 발견: {elem_text}")
                                            break
                                if actual_fund_name != "찾을 수 없음":
                                    break
                        except:
                            pass
                        
                        # 방법 2: 테이블 내부 caption 확인
                        if actual_fund_name == "찾을 수 없음":
                            try:
                                caption = main_table.find_element(By.TAG_NAME, "caption")
                                caption_text = caption.text.strip()
                                if "형" in caption_text:
                                    fund_keywords = ["혼합성장형", "글로벌주식형", "주식형", "채권형", "안정형", "MMF형", "성장형"]
                                    for keyword in fund_keywords:
                                        if keyword in caption_text:
                                            actual_fund_name = keyword
                                            print(f"      → 테이블 caption에서 발견: {caption_text}")
                                            break
                            except:
                                pass
                        
                        # 방법 3: 테이블의 부모 컨테이너에서 찾기
                        if actual_fund_name == "찾을 수 없음":
                            try:
                                parent = main_table.find_element(By.XPATH, "..")
                                # 부모 내의 strong, b, h3 등 강조 요소에서 찾기
                                emphasis_elements = parent.find_elements(By.XPATH, ".//*[self::strong or self::b or self::h3 or self::h4]")
                                for elem in emphasis_elements:
                                    elem_text = elem.text.strip()
                                    if "형" in elem_text and len(elem_text) < 30:
                                        fund_keywords = ["혼합성장형", "글로벌주식형", "주식형", "채권형", "안정형", "MMF형", "성장형"]
                                        for keyword in fund_keywords:
                                            if keyword in elem_text:
                                                actual_fund_name = keyword
                                                print(f"      → 테이블 부모 컨테이너에서 발견: {elem_text}")
                                                break
                                    if actual_fund_name != "찾을 수 없음":
                                        break
                            except:
                                pass
                    
                    # 방법 4: JavaScript로 테이블 바로 위의 텍스트 노드 읽기
                    if actual_fund_name == "찾을 수 없음":
                        try:
                            table_context = driver.execute_script("""
                                var tables = document.getElementsByTagName('table');
                                if (tables.length > 0) {
                                    var table = tables[0];
                                    var prev = table.previousElementSibling;
                                    var texts = [];
                                    for (var i = 0; i < 5 && prev; i++) {
                                        if (prev.textContent) {
                                            texts.push(prev.textContent.trim());
                                        }
                                        prev = prev.previousElementSibling;
                                    }
                                    return texts.join(' | ');
                                }
                                return '';
                            """)
                            
                            if table_context:
                                print(f"      → 테이블 주변 텍스트: {table_context[:100]}")
                                fund_keywords = ["혼합성장형", "글로벌주식형", "주식형", "채권형", "안정형", "MMF형", "성장형"]
                                for keyword in fund_keywords:
                                    if keyword in table_context:
                                        actual_fund_name = keyword
                                        break
                        except:
                            pass
                    
                except Exception as e:
                    print(f"      ⚠ 펀드명 추출 실패: {e}")
                
                print(f"      📌 테이블 상단 펀드명: {actual_fund_name}")
                
                # 7. 스크린샷 저장 (매 시도마다)
                # 파일명 안전하게 만들기
                safe_label = re.sub(r'[^\w\s-]', '', checkbox_label)[:15]
                safe_fund = re.sub(r'[^\w\s-]', '', actual_fund_name)[:15]
                screenshot_name = f"시도{attempt_num:02d}_원래{idx:02d}_{safe_label}_{safe_fund}.png"
                screenshot_name = screenshot_name.replace(' ', '_')
                
                driver.save_screenshot(screenshot_name)
                print(f"      📸 스크린샷: {screenshot_name}")
                
                # 8. 결과 확인: 실제 펀드명이 "혼합성장형"인지 체크
                print(f"\n      [판정]")
                
                # 중요: actual_fund_name으로 판단! (페이지 소스가 아님)
                is_mixed_growth = ("혼합성장형" in actual_fund_name)
                
                if is_mixed_growth:
                    print(f"\n   ✅✅✅ 성공! '혼합성장형' 발견! ✅✅✅")
                    print(f"   → 체크박스 번호: {idx}")
                    print(f"   → 체크박스 라벨: {checkbox_label}")
                    print(f"   → 검색 결과 펀드명: {actual_fund_name}")
                    
                    mixed_found = True
                    found_fund_name = actual_fund_name  # 펀드명 저장
                    
                    # 성공 스크린샷 별도 저장
                    success_screenshot = f"SUCCESS_혼합성장형_발견.png"
                    driver.save_screenshot(success_screenshot)
                    print(f"      📸 성공 스크린샷 저장: {success_screenshot}")
                    
                    print(f"\n   🎯 혼합성장형을 찾았으므로 다른 펀드는 검색하지 않습니다.")
                    print(f"   → 바로 데이터 수집으로 이동합니다...")
                    break  # 루프 종료!
                else:
                    # 다른 펀드 유형임
                    print(f"      ✗ '혼합성장형'이 아닙니다!")
                    print(f"      → 실제 펀드명: {actual_fund_name}")
                    print(f"      → 체크박스 라벨: {checkbox_label}")
                    print(f"      → 스크린샷 확인: {screenshot_name}")
                    print(f"      → 다음 체크박스 시도...")
                    
            except Exception as e:
                print(f"      ✗ 시도 {attempt_num} (원래 {idx}번 체크박스) 테스트 실패: {e}")
                # 실패 스크린샷
                try:
                    error_screenshot = f"ERROR_시도{attempt_num:02d}.png"
                    driver.save_screenshot(error_screenshot)
                    print(f"      📸 에러 스크린샷: {error_screenshot}")
                except:
                    pass
                continue
        
        # 최종 결과
        print("\n" + "="*80)
        print("[8단계] 펀드 선택 결과")
        print("="*80)
        if mixed_found:
            print("   ✅✅✅ '혼합성장형' 선택 완료! ✅✅✅")
            print("   → 성공 스크린샷: SUCCESS_혼합성장형_발견.png")
        else:
            print("   ❌❌❌ 경고: '혼합성장형'을 찾지 못했습니다! ❌❌❌")
            print(f"   → 총 {len(checkbox_info)}개의 체크박스를 모두 시도했습니다.")
            print("   → 현재 폴더에 저장된 스크린샷들을 확인하세요:")
            print(f"      시도*.png 파일들")
            print("\n   ⚠ 마지막 선택된 펀드로 데이터 수집을 시도합니다...")
            driver.save_screenshot("warning_mixed_not_found.png")
            print("   → 최종 스크린샷: warning_mixed_not_found.png")
        
        print("\n[9단계] 데이터 수집 중...")
        try:
            table = wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
            rows = table.find_elements(By.TAG_NAME, "tr")
            
            all_data = []
            headers = []
            
            print(f"   → 총 {len(rows)}개의 테이블 행 발견")
            
            # 헤더와 데이터 분리 수집
            for row_idx, row in enumerate(rows):
                # 모든 셀 찾기 (th와 td 모두)
                all_cells = row.find_elements(By.XPATH, ".//th | .//td")
                
                if not all_cells:
                    continue
                
                cell_texts = [cell.text.strip() for cell in all_cells]
                
                # 디버깅: 처음 5개 행 출력
                if row_idx < 5:
                    print(f"   → 행 {row_idx}: {cell_texts}")
                
                # 헤더 찾기: "날짜"나 "기준가" 같은 단어가 있으면 헤더로 간주
                if not headers and cell_texts:
                    if any(keyword in ''.join(cell_texts) for keyword in ['날짜', '일자', '기준가', '가격']):
                        headers = cell_texts
                        print(f"   ✅ 헤더 발견 (행 {row_idx}): {headers}")
                        continue
                
                # 데이터 행: 빈 셀이 아니고 숫자나 날짜 형식이 있으면 데이터로 간주
                if cell_texts and any(cell_texts):
                    # 헤더가 이미 설정되었거나, 날짜 형식(YYYY-MM-DD 또는 숫자)이 있으면 데이터
                    if headers or any(c for c in cell_texts if c and (c.replace('-', '').replace('.', '').replace(',', '').isdigit() or '-' in c)):
                        all_data.append(cell_texts)
            
            print(f"   → 수집된 데이터 행: {len(all_data)}개")
            print(f"   → 수집된 헤더: {headers}")
            
            if all_data:
                # 데이터 샘플 출력
                print(f"   → 첫 번째 데이터 행: {all_data[0] if all_data else 'None'}")
                
                # 헤더가 없거나 길이가 안 맞으면 기본 헤더 사용
                if not headers:
                    print(f"   ⚠ 헤더를 찾지 못했습니다. 기본 헤더 사용")
                    if len(all_data[0]) == 2:
                        headers = ['날짜', '기준가(원)']
                    else:
                        headers = [f'컬럼{i+1}' for i in range(len(all_data[0]))]
                    print(f"   → 기본 헤더: {headers}")
                elif len(headers) != len(all_data[0]):
                    print(f"   ⚠ 헤더 길이({len(headers)})와 데이터 길이({len(all_data[0])}) 불일치")
                    # 데이터 길이에 맞춰 헤더 조정
                    if len(all_data[0]) == 2:
                        headers = ['날짜', '기준가(원)']
                    else:
                        headers = headers[:len(all_data[0])] if len(headers) > len(all_data[0]) else headers + [f'컬럼{i+1}' for i in range(len(headers), len(all_data[0]))]
                    print(f"   → 조정된 헤더: {headers}")
                
                # DataFrame 생성
                df = pd.DataFrame(all_data, columns=headers)
                
                # 수집 날짜 추가
                collection_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                df['수집일시'] = collection_date
                
                # 엑셀 저장 (펀드명 + 타임스탬프)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'메트라이프_혼합성장형_기준가_{timestamp}.xlsx'
                save_path = os.path.join(os.getcwd(), filename)
                fund_name_display = found_fund_name if found_fund_name else "혼합성장형"
                
                try:
                    # openpyxl을 사용하여 엑셀에 펀드명 추가
                    from openpyxl import Workbook
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    from openpyxl.styles import Font, Alignment
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "기준가 데이터"
                    
                    # 첫 행: 펀드명 (굵게, 크게)
                    ws.append([f"펀드명: {fund_name_display}"])
                    ws['A1'].font = Font(bold=True, size=14)
                    
                    # 두 번째 행: 수집일시
                    ws.append([f"수집일시: {collection_date}"])
                    ws['A2'].font = Font(bold=True, size=11)
                    
                    ws.append([])  # 빈 줄
                    
                    # 데이터프레임 추가 (헤더 굵게)
                    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), start=4):
                        ws.append(r)
                        if r_idx == 4:  # 헤더 행
                            for cell in ws[r_idx]:
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal='center')
                    
                    # 컬럼 너비 자동 조정
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    wb.save(save_path)
                    
                except ImportError:
                    # openpyxl이 없으면 기본 pandas to_excel 사용
                    print("   ⚠ openpyxl이 없어 기본 형식으로 저장합니다.")
                    # 펀드명과 수집일시를 데이터프레임 첫 행에 추가
                    df_with_info = pd.DataFrame([
                        [f"펀드명: {fund_name_display}"] + [''] * (len(df.columns) - 1),
                        [f"수집일시: {collection_date}"] + [''] * (len(df.columns) - 1),
                        [''] * len(df.columns)
                    ], columns=df.columns)
                    df_final = pd.concat([df_with_info, df], ignore_index=True)
                    df_final.to_excel(save_path, index=False)
                    df = df_final  # 미리보기용
                
                print(f"\n[10단계] ✅ 엑셀 저장 완료: {save_path}")
                print(f"   → 펀드명: {fund_name_display}")
                print(f"   → 수집일시: {collection_date}")
                print(f"   → 수집된 데이터: {len(df)}개 행")
                print("\n데이터 미리보기:")
                print(df.head(10))
            else:
                print("   ❌ 수집된 데이터가 없습니다.")
        except Exception as e:
            print(f"   ❌ 데이터 수집 실패: {e}")
            driver.save_screenshot("error_data_collection.png")
            print("   스크린샷 저장: error_data_collection.png")

        print("\n" + "="*80)
        print("✅ 모든 작업이 완료되었습니다!")
        print("="*80)
        print("\n5초 후 브라우저가 닫힙니다...")
        time.sleep(5)
        
    except Exception as e:
        print("\n" + "="*80)
        print("❌ 크리티컬 오류 발생!")
        print("="*80)
        print(f"\n오류 내용: {e}")
        screenshot_name = f"error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        driver.save_screenshot(screenshot_name)
        print(f"\n스크린샷 저장됨: {screenshot_name}")
        print("\n오류 내용을 확인하신 후 [Enter]를 눌러주세요...")
        input(">>> [Enter] 입력: ")
    finally:
        print("\n브라우저를 종료합니다...")
        driver.quit()

if __name__ == "__main__":
    main()
